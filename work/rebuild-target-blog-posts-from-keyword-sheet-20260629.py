#!/usr/bin/env python3
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_SPREADSHEET_ID = "1Cgn-CFMEobWpwhdmrJxVKgVLMblGQ_JEX4xVpXS4_ZA"
SOURCE_SHEETS = [
    ("월보장 시트", "0"),
    ("건바이 노출 시트", "2026584527"),
]
DEFAULT_PUBLIC_POSTS_JSON = (
    "outputs/blog-published-ranks/target-blog-public-posts-20260629105240.json"
)
OUTPUT_DIR = Path("outputs/blog-published-ranks")
HEADERS = [
    "순위",
    "글",
    "링크",
    "업체명",
    "키워드",
    "발행일",
    "글번호",
    "포함사유",
    "동일키워드발행수",
    "키워드출처",
    "시트중복수",
]


@dataclass(frozen=True)
class KeywordEntry:
    source_sheet: str
    source_row: int
    company: str
    keyword: str
    norm_keyword: str


def normalize(value: str) -> str:
    return re.sub(r"[\s\(\)\[\]\{\}'\"`.,:;!?|/_\\\-]+", "", value or "").lower()


def fetch_csv_rows(gid: str) -> list[list[str]]:
    url = (
        f"https://docs.google.com/spreadsheets/d/{SOURCE_SPREADSHEET_ID}"
        f"/export?format=csv&gid={gid}"
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=30) as response:
        text = response.read().decode("utf-8-sig", errors="replace")
    return list(csv.reader(text.splitlines()))


def load_keyword_entries() -> list[KeywordEntry]:
    entries: list[KeywordEntry] = []
    seen_source_rows: set[tuple[str, int, str]] = set()

    for source_sheet, gid in SOURCE_SHEETS:
        last_company = ""
        for row_index, row in enumerate(fetch_csv_rows(gid), start=1):
            company = (row[0] if len(row) > 0 else "").strip()
            keyword = (row[1] if len(row) > 1 else "").strip()

            if company and company not in {"업체명"}:
                last_company = company

            if not keyword or keyword == "키워드":
                continue

            norm_keyword = normalize(keyword)
            if not norm_keyword:
                continue

            key = (source_sheet, row_index, norm_keyword)
            if key in seen_source_rows:
                continue
            seen_source_rows.add(key)

            entries.append(
                KeywordEntry(
                    source_sheet=source_sheet,
                    source_row=row_index,
                    company=company or last_company,
                    keyword=keyword,
                    norm_keyword=norm_keyword,
                )
            )

    return entries


def parse_generated_date(value: str) -> date:
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", value or "")
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return datetime.now().date()


def parse_post_date(value: str, reference: date) -> date | None:
    raw = (value or "").strip()
    match = re.search(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\.", raw)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    if re.search(r"\d+\s*(초|분|시간)\s*전", raw) or raw in {"방금 전", "오늘"}:
        return reference
    if raw == "어제":
        return reference - timedelta(days=1)
    return None


def summarize_companies(entries: list[KeywordEntry]) -> str:
    companies = []
    seen = set()
    for entry in entries:
        company = entry.company.strip()
        if not company or company in seen:
            continue
        seen.add(company)
        companies.append(company)
    if len(companies) <= 5:
        return ", ".join(companies)
    return f"{', '.join(companies[:5])} 외 {len(companies) - 5}"


def choose_keyword_display(entries: list[KeywordEntry]) -> str:
    keywords = []
    seen = set()
    for entry in entries:
        keyword = entry.keyword.strip()
        compact = normalize(keyword)
        if compact in seen:
            continue
        seen.add(compact)
        keywords.append(keyword)
    return " / ".join(keywords[:3]) if keywords else ""


def choose_source_display(entries: list[KeywordEntry]) -> str:
    sources = []
    seen = set()
    for entry in entries:
        source = f"{entry.source_sheet} R{entry.source_row}"
        if source in seen:
            continue
        seen.add(source)
        sources.append(source)
    if len(sources) <= 3:
        return ", ".join(sources)
    return f"{', '.join(sources[:3])} 외 {len(sources) - 3}"


def build_keyword_index(entries: list[KeywordEntry]) -> dict[str, list[KeywordEntry]]:
    grouped: dict[str, list[KeywordEntry]] = defaultdict(list)
    for entry in entries:
        grouped[entry.norm_keyword].append(entry)
    return dict(grouped)


def match_keywords(title: str, keyword_index: dict[str, list[KeywordEntry]]) -> list[str]:
    norm_title = normalize(title)
    matches = [keyword for keyword in keyword_index if keyword and keyword in norm_title]
    if not matches:
        return []

    matches.sort(key=lambda value: (-len(value), value))
    selected: list[str] = []
    for keyword in matches:
        if any(keyword in existing for existing in selected):
            continue
        selected.append(keyword)
    return selected


def make_cell(value: Any) -> dict[str, Any]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return {"userEnteredValue": {"numberValue": value}}
    return {"userEnteredValue": {"stringValue": "" if value is None else str(value)}}


def write_xlsx(results: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for result in results:
        worksheet = workbook.create_sheet(result["sheetName"])
        worksheet.append(HEADERS)
        for row in result["rows"]:
            worksheet.append([row.get(header, "") for header in HEADERS])

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(HEADERS))}{max(1, len(result['rows']) + 1)}"
        )
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = {
            "A": 8,
            "B": 52,
            "C": 44,
            "D": 28,
            "E": 22,
            "F": 16,
            "G": 18,
            "H": 22,
            "I": 18,
            "J": 28,
            "K": 12,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)


def main() -> None:
    public_json_path = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PUBLIC_POSTS_JSON)
    payload = json.loads(public_json_path.read_text(encoding="utf-8"))
    reference_date = parse_generated_date(payload.get("generatedAtKst", ""))
    cutoff = date(reference_date.year, reference_date.month, reference_date.day)
    cutoff = cutoff.replace(month=cutoff.month - 3) if cutoff.month > 3 else cutoff.replace(
        year=cutoff.year - 1, month=cutoff.month + 9
    )

    keyword_entries = load_keyword_entries()
    keyword_index = build_keyword_index(keyword_entries)

    matched_by_sheet: list[dict[str, Any]] = []
    all_summary = []

    for result in payload["results"]:
        raw_matches: list[dict[str, Any]] = []
        duplicate_counter: Counter[str] = Counter()

        for post in result["rows"]:
            matched_keywords = match_keywords(post.get("글", ""), keyword_index)
            for norm_keyword in matched_keywords:
                duplicate_counter[norm_keyword] += 1
                raw_matches.append({"post": post, "normKeyword": norm_keyword})

        rows = []
        for item in raw_matches:
            post = item["post"]
            norm_keyword = item["normKeyword"]
            entries = keyword_index[norm_keyword]
            post_date = parse_post_date(post.get("발행일", ""), reference_date)
            is_recent = post_date is not None and post_date >= cutoff
            duplicate_count = duplicate_counter[norm_keyword]
            is_duplicate = duplicate_count >= 2

            if not is_recent and not is_duplicate:
                continue

            if is_recent and is_duplicate:
                reason = "최근3개월+동일키워드"
            elif is_recent:
                reason = "최근3개월"
            else:
                reason = "동일키워드"

            rows.append(
                {
                    "순위": post.get("순위", ""),
                    "글": post.get("글", ""),
                    "링크": post.get("링크", ""),
                    "업체명": summarize_companies(entries),
                    "키워드": choose_keyword_display(entries),
                    "발행일": post.get("발행일", ""),
                    "글번호": post.get("글번호", ""),
                    "포함사유": reason,
                    "동일키워드발행수": duplicate_count,
                    "키워드출처": choose_source_display(entries),
                    "시트중복수": len(entries),
                }
            )

        rows.sort(key=lambda row: (int(row["순위"] or 999999), row["키워드"]))
        matched_by_sheet.append(
            {
                "sheetName": result["sheetName"],
                "displayName": result.get("displayName", ""),
                "blogId": result.get("blogId", ""),
                "rows": rows,
            }
        )
        all_summary.append(
            {
                "sheetName": result["sheetName"],
                "blogId": result.get("blogId", ""),
                "rows": len(rows),
                "matchedBeforeFilter": len(raw_matches),
                "uniqueMatchedKeywords": len({item["normKeyword"] for item in raw_matches}),
            }
        )

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUTPUT_DIR / f"target-blog-posts-keyword-sheet-3months-duplicates-{timestamp}.json"
    xlsx_path = OUTPUT_DIR / f"target-blog-posts-keyword-sheet-3months-duplicates-{timestamp}.xlsx"

    output = {
        "generatedAtKst": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sourceKeywordSpreadsheetId": SOURCE_SPREADSHEET_ID,
        "sourceKeywordSheets": [sheet for sheet, _gid in SOURCE_SHEETS],
        "publicPostsJson": str(public_json_path),
        "cutoffDate": cutoff.isoformat(),
        "rankingBasis": payload.get("rankingBasis", "각 블로그 공개 글 목록 최신순 순위"),
        "filterBasis": "최근 3개월 글 전체 + 최근 3개월 이전 동일 키워드 2건 이상 발행글",
        "keywordBasis": "루트컴퍼니 전체현황 (신규) 월보장/건바이 시트 B열 키워드",
        "headers": HEADERS,
        "results": matched_by_sheet,
        "summary": all_summary,
        "keywordEntryCount": len(keyword_entries),
        "uniqueKeywordCount": len(keyword_index),
    }

    json_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(matched_by_sheet, xlsx_path)

    print(
        json.dumps(
            {
                "jsonPath": str(json_path),
                "xlsxPath": str(xlsx_path),
                "cutoffDate": cutoff.isoformat(),
                "keywordEntryCount": len(keyword_entries),
                "uniqueKeywordCount": len(keyword_index),
                "summary": all_summary,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
