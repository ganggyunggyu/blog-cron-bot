#!/usr/bin/env python3
import csv
import io
import json
import os
import re
import time
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_SHEET_ID = "1Cgn-CFMEobWpwhdmrJxVKgVLMblGQ_JEX4xVpXS4_ZA"
SOURCE_TABS = [
    ("월보장 시트", "0"),
    ("건바이 노출 시트", "2026584527"),
]
PUBLIC_POSTS_JSON = Path(os.environ.get("PUBLIC_POSTS_JSON", "outputs/blog-published-ranks/target-blog-public-posts-20260629105240.json"))
RANK_JSON = Path(os.environ["RANK_JSON"]) if os.environ.get("RANK_JSON") else None
OUT_BASE = os.environ.get("OUT_BASE", "target-blog-posts-keyword-sheet-mar-jun-all-with-rank-20260629")
OUT_DIR = Path("outputs/blog-published-ranks")
TSV_DIR = Path(os.environ.get("TSV_DIR", "work/mar-jun-all-with-rank-tsv-20260629"))
START_DATE = datetime.strptime(os.environ.get("START_DATE", "2026-03-01"), "%Y-%m-%d")
END_DATE = datetime.strptime(os.environ.get("END_DATE", "2026-06-29"), "%Y-%m-%d")

OUTPUT_COLUMNS = ["링크", "업체명", "키워드", "발행일", "동일키워드발행수", "검색순위"]
BAD_KEYWORDS = {
    "",
    "o",
    "x",
    "키워드",
    "업체명",
    "노출",
    "as",
    "as여부",
    "견적",
    "비고",
    "링크",
    "인기주제",
}
GENERIC_COMPANY_TOKENS = {
    "본점",
    "직영점",
    "분점",
    "신규",
    "맛집",
    "카페",
    "식당",
    "레스토랑",
    "이자카야",
    "오마카세",
    "고기집",
    "횟집",
    "술집",
    "점",
    "역",
    "요리",
    "구이",
    "오리구이",
}
COMPANY_SUFFIXES = (
    "오리구이",
    "돌판요리",
    "요리",
    "구이",
    "본점",
    "직영점",
    "분점",
    "점",
)
BAD_COMPANY_FRAGMENTS = (
    "맛집",
    "맛",
    "곳만",
    "광고",
    "후기",
    "혼밥",
    "추천",
    "거리뷰",
    "개의",
)


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").lower()
    return re.sub(r"[^0-9a-z가-힣]+", "", value)


def normalize_keyword(value: str) -> str:
    return normalize_text(value)


def visible(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def fetch_text(url: str, timeout: int = 30) -> str:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            ),
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as res:
        return res.read().decode("utf-8-sig", errors="replace")


def load_source_keywords():
    entries = []
    by_keyword = defaultdict(list)

    for tab_name, gid in SOURCE_TABS:
        url = f"https://docs.google.com/spreadsheets/d/{SOURCE_SHEET_ID}/export?format=csv&gid={gid}"
        text = fetch_text(url)
        rows = list(csv.reader(io.StringIO(text)))
        current_company = ""

        for row_no, row in enumerate(rows, start=1):
            row = [visible(cell) for cell in row]
            if len(row) < 2:
                continue

            company = row[0] or current_company
            keyword = row[1]

            if row[0]:
                current_company = row[0]

            keyword_norm = normalize_keyword(keyword)
            if keyword_norm in BAD_KEYWORDS or len(keyword_norm) < 2:
                continue
            if normalize_text(company) in BAD_KEYWORDS or not normalize_text(company):
                continue
            if keyword_norm == normalize_text(company):
                continue

            entry = {
                "company": company,
                "companyNorm": normalize_text(company),
                "keyword": keyword,
                "keywordNorm": keyword_norm,
                "sourceTab": tab_name,
                "rowNo": row_no,
            }
            entries.append(entry)
            by_keyword[keyword_norm].append(entry)

    for keyword_norm, candidates in list(by_keyword.items()):
        seen = set()
        deduped = []
        for item in candidates:
            key = (item["companyNorm"], item["keywordNorm"])
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        by_keyword[keyword_norm] = deduped

    return entries, by_keyword


def parse_reference_date(value: str) -> datetime:
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y. %m. %d.", "%Y. %-m. %-d."):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return datetime.now()


def parse_post_date(value: str, reference: datetime) -> datetime | None:
    value = visible(value)
    match = re.fullmatch(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\.", value)
    if match:
        y, m, d = map(int, match.groups())
        return datetime(y, m, d)
    match = re.fullmatch(r"(\d+)\s*시간 전", value)
    if match:
        return reference - timedelta(hours=int(match.group(1)))
    match = re.fullmatch(r"(\d+)\s*분 전", value)
    if match:
        return reference - timedelta(minutes=int(match.group(1)))
    match = re.fullmatch(r"(\d+)\s*일 전", value)
    if match:
        return reference - timedelta(days=int(match.group(1)))
    return None


@lru_cache(maxsize=4096)
def keyword_pattern(keyword: str) -> re.Pattern:
    compact = re.sub(r"\s+", "", visible(keyword))
    separators = r"[\s·ㆍ|ㅣ_\-]*"
    body = separators.join(re.escape(char) for char in compact)
    return re.compile(rf"(?<![0-9A-Za-z가-힣]){body}(?![0-9A-Za-z가-힣])", re.IGNORECASE)


def title_has_keyword(title: str, keyword: str) -> bool:
    return bool(keyword_pattern(keyword).search(visible(title)))


def choose_keyword(title: str, keyword_entries):
    title_norm = normalize_keyword(title)
    matches = []
    for item in keyword_entries:
        kw_norm = item["keywordNorm"]
        if kw_norm and kw_norm in title_norm and title_has_keyword(title, item["keyword"]):
            matches.append(item)
    if not matches:
        return None
    matches.sort(key=lambda item: (len(item["keywordNorm"]), len(item["keyword"])), reverse=True)
    return matches[0]


def fetch_post_plain_text(blog_id: str, log_no: str) -> str:
    params = urllib.parse.urlencode(
        {
            "blogId": blog_id,
            "logNo": log_no,
            "redirect": "Dlog",
            "widgetTypeCall": "true",
            "directAccess": "false",
        }
    )
    url = f"https://blog.naver.com/PostView.naver?{params}"
    html = fetch_text(url, timeout=20)
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(" ", strip=True)


def company_tokens(company: str):
    rough = re.findall(r"[0-9A-Za-z가-힣]+", unicodedata.normalize("NFKC", company or ""))
    tokens = []
    for token in rough:
        token_norm = normalize_text(token)
        if not token_norm or token_norm in GENERIC_COMPANY_TOKENS:
            continue
        if len(token_norm) < 2:
            continue
        tokens.append(token_norm)

    full = normalize_text(company)
    for suffix in COMPANY_SUFFIXES:
        suffix_norm = normalize_text(suffix)
        if full.endswith(suffix_norm) and len(full) - len(suffix_norm) >= 2:
            tokens.append(full[: -len(suffix_norm)])

    trimmed = re.sub(
        r"(광안리|강남|홍대|건대|송도|동탄|수원|청주|율량|일산|안산|예산|강릉|용산|동대구|신논현|방이|부천|김해|남천|사당|대전|마곡|종각|구디|구로)?(본점|직영점|점)$",
        "",
        full,
    )
    if trimmed and len(trimmed) >= 2:
        tokens.append(trimmed)

    deduped = []
    seen = set()
    for token in sorted(tokens, key=len, reverse=True):
        if token not in seen:
            seen.add(token)
            deduped.append(token)
    return deduped


def clean_company_name(value: str) -> str:
    value = visible(value)
    value = re.sub(r"^(?:추천|강추|위치|기본|정보|매장|가게|식당)\s+", "", value)
    value = re.sub(r"\s*(?:외관|분위기|위치|영업|기본|메뉴|내부|자리|공간|느낌).*$", "", value)
    value = re.sub(r"의$", "", value)
    value = re.sub(r"(?:이라는 곳|라는 곳|이라는|라는|입니다|이에요|였어요|이고요|인데요|인데|으로|에서|은|는|이|가|을|를|집)$", "", value)
    value = re.sub(r"으$", "", value)
    value = re.sub(r"(?:\s*추천|\s*기본정보|\s*기본 정보|\s*위치와.*|\s*위치.*|\s*영업.*)$", "", value)
    return visible(value)


def is_plausible_company_name(value: str) -> bool:
    norm = normalize_text(value)
    if not (2 <= len(norm) <= 30):
        return False
    if norm in GENERIC_COMPANY_TOKENS:
        return False
    if norm.endswith(("맛", "맛집", "곳만", "후기")):
        return False
    if any(fragment in norm for fragment in ("광고", "내돈내산", "추천만", "후기만", "거리뷰", "개의글")):
        return False
    if norm in BAD_COMPANY_FRAGMENTS:
        return False
    return True


def extract_name_around_token(text: str, token_norm: str) -> str:
    if not token_norm:
        return ""
    candidates = re.findall(r"[0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,2}", text or "")
    bad_fragments = ("맛집", "추천", "후기", "한그릇", "30cm", "붓카케", "시그니", "이번엔", "입니다")
    for raw in candidates:
        raw_clean = clean_company_name(raw)
        raw_norm = normalize_text(raw_clean)
        if not raw_norm.startswith(token_norm):
            continue
        if len(raw_norm) > len(token_norm) + 10:
            continue
        if any(fragment in raw_norm for fragment in bad_fragments):
            continue
        if len(raw_norm) < len(token_norm):
            continue
        if not is_plausible_company_name(raw_clean):
            continue
        return raw_clean
    return ""


def extract_company_from_title(title: str, keyword: str) -> str:
    title = visible(title)
    keyword = visible(keyword)

    direct_patterns = [
        r"(?:가는|가본|다녀온)\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,2})$",
        r"맛집\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,2})$",
        r"[ㅣ|]\s*([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,2})$",
    ]
    for pattern in direct_patterns:
        match = re.search(pattern, title)
        if match:
            candidate = clean_company_name(match.group(1))
            if is_plausible_company_name(candidate):
                return candidate

    cleaned = title
    for pattern in [
        r"\b추천\b.*$",
        r"\b후기\b.*$",
    ]:
        cleaned = re.sub(pattern, "", cleaned).strip()

    if "맛집" in cleaned:
        tail = cleaned.rsplit("맛집", 1)[-1].strip()
        tail = clean_company_name(tail)
        tail_norm = normalize_text(tail)
        if is_plausible_company_name(tail) and not re.search(r"(좋은곳|좋은|추천)$", tail_norm):
            return tail

    keywordless = cleaned.replace(keyword, "").strip() if keyword else cleaned
    words = keywordless.split()
    if words:
        tail = clean_company_name(" ".join(words[-2:]))
        if is_plausible_company_name(tail):
            return tail
    return ""


def extract_company_from_body(body: str) -> str:
    body = visible(body)
    patterns = [
        r"([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})\s+주소\s*[:：]",
        r"([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})\s+위치와",
        r"그런데\s+([0-9A-Za-z가-힣]+)(?:은|는)\s",
        r"찾다가\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})(?:으로|로)\s+발길",
        r"소개해드릴(?:게요|께요|까\s*해요)\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,2})",
        r"\d+\.\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})\s",
        r"자리한\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})(?:이|가|은|는)\s+그곳",
        r"다녀온\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})(?:이|가|은|는)\s",
        r"\d+\.\s+([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,1})\s+(?:기본\s*정보|위치와\s*영업|위치\s*정보)",
        r"([0-9A-Za-z가-힣]+(?:\s+[0-9A-Za-z가-힣]+){0,2})(?:이라는 곳|라는 곳)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, body):
            candidate = clean_company_name(match.group(1))
            if is_plausible_company_name(candidate):
                return candidate
    return ""


def match_company(candidates, title: str, body: str):
    text_norm = normalize_text(f"{title} {body}")
    title_norm = normalize_text(title)
    best = None

    for item in candidates:
        company = item["company"]
        company_norm = item["companyNorm"]
        if not company_norm:
            continue

        score = 0
        reason = ""
        actual = company
        if company_norm in text_norm:
            score = 10000 + len(company_norm)
            reason = "full"
        else:
            tokens = company_tokens(company)
            matched = [token for token in tokens if token in text_norm]
            if matched:
                longest = max(matched, key=len)
                extracted = extract_name_around_token(f"{title} {body}", longest)
                if normalize_text(extracted) == "라성돌판오리구이":
                    actual = extracted
                score = 5000 + len(longest) * 10 + len(matched)
                reason = f"token:{longest}"
                if longest in title_norm:
                    score += 500

        if score and (best is None or score > best["score"]):
            best = {"company": actual or company, "score": score, "reason": reason}

    if best:
        return best["company"], best["reason"]

    body_company = extract_company_from_body(body)
    if body_company:
        return body_company, "body"

    title_company = extract_company_from_title(title, "")
    if title_company:
        return title_company, "title"

    if len(candidates) == 1:
        return candidates[0]["company"], "single-source-fallback"

    return "업체명 확인필요", "missing"


def make_cell(value):
    if isinstance(value, int):
        return value
    return value or ""


def write_xlsx(results, path: Path):
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for blog in results:
        ws = wb.create_sheet(blog["sheetName"])
        ws.append(OUTPUT_COLUMNS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in blog["rows"]:
            ws.append([make_cell(row.get(col)) for col in OUTPUT_COLUMNS])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:E{max(1, ws.max_row)}"
        widths = [52, 24, 20, 14, 16, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(path)


def write_tsv(blog, path: Path):
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(OUTPUT_COLUMNS)
        for row in blog["rows"]:
            writer.writerow([row.get(col, "") for col in OUTPUT_COLUMNS])


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    TSV_DIR.mkdir(parents=True, exist_ok=True)

    entries, by_keyword = load_source_keywords()
    keyword_entries = []
    seen_keyword_company = set()
    for item in entries:
        key = (item["keywordNorm"], item["companyNorm"])
        if key in seen_keyword_company:
            continue
        seen_keyword_company.add(key)
        keyword_entries.append(item)

    with PUBLIC_POSTS_JSON.open(encoding="utf-8") as f:
        public_data = json.load(f)
    rank_by_link = {}
    if RANK_JSON and RANK_JSON.exists():
        with RANK_JSON.open(encoding="utf-8") as f:
            rank_data = json.load(f)
        rank_by_link = rank_data.get("rankByLink", {})

    reference = parse_reference_date(public_data.get("generatedAtKst") or "")
    results = []
    fetch_cache = {}
    diagnostics = {
        "sourceKeywordRows": len(entries),
        "sourceUniqueKeywords": len(by_keyword),
        "startDate": START_DATE.strftime("%Y-%m-%d"),
        "endDate": END_DATE.strftime("%Y-%m-%d"),
        "generatedAtKst": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "companyMissingCount": 0,
        "companyMissingSamples": [],
        "removedDuplicateLinks": {},
    }

    for blog in public_data["results"]:
        provisional = []
        link_seen = set()
        duplicate_removed = 0

        for post in blog["rows"]:
            post_dt = parse_post_date(post.get("발행일", ""), reference)
            if post_dt is None or post_dt.date() < START_DATE.date() or post_dt.date() > END_DATE.date():
                continue

            chosen = choose_keyword(post.get("글", ""), keyword_entries)
            if not chosen:
                continue

            link = visible(post.get("링크", ""))
            if not link:
                continue
            if link in link_seen:
                duplicate_removed += 1
                continue
            link_seen.add(link)

            provisional.append(
                {
                    "링크": link,
                    "키워드": chosen["keyword"],
                    "keywordNorm": chosen["keywordNorm"],
                    "발행일": post.get("발행일", ""),
                    "postDateSort": post_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "blogId": post.get("블로그ID", blog.get("blogId")),
                    "logNo": post.get("글번호", ""),
                    "title": post.get("글", ""),
                }
            )

        counts = Counter(row["keywordNorm"] for row in provisional)
        rows = provisional

        for row in rows:
            cache_key = (row["blogId"], row["logNo"])
            if cache_key not in fetch_cache:
                try:
                    fetch_cache[cache_key] = fetch_post_plain_text(*cache_key)
                    time.sleep(0.08)
                except Exception as exc:
                    fetch_cache[cache_key] = ""
                    row["fetchError"] = str(exc)
            candidates = by_keyword.get(row["keywordNorm"], [])
            company, reason = match_company(candidates, row["title"], fetch_cache[cache_key])
            row["업체명"] = company
            row["companyMatchReason"] = reason
            row["동일키워드발행수"] = counts[row["keywordNorm"]]
            rank_value = rank_by_link.get(row["링크"]) or rank_by_link.get(row["링크"].replace("https://blog.naver.com/", "https://m.blog.naver.com/"))
            row["검색순위"] = rank_value or ""
            if company == "업체명 확인필요":
                diagnostics["companyMissingCount"] += 1
                if len(diagnostics["companyMissingSamples"]) < 20:
                    diagnostics["companyMissingSamples"].append(
                        {
                            "sheetName": blog["sheetName"],
                            "keyword": row["키워드"],
                            "title": row["title"],
                            "link": row["링크"],
                            "candidateCompanies": [item["company"] for item in candidates[:20]],
                        }
                    )

        rows.sort(key=lambda row: (row["keywordNorm"], -datetime.strptime(row["postDateSort"], "%Y-%m-%d %H:%M:%S").timestamp(), row["링크"]))
        clean_rows = [{col: row.get(col, "") for col in OUTPUT_COLUMNS} for row in rows]
        results.append(
            {
                "sheetName": blog["sheetName"],
                "displayName": blog.get("displayName", ""),
                "blogId": blog.get("blogId", ""),
                "rows": clean_rows,
            }
        )
        diagnostics["removedDuplicateLinks"][blog["sheetName"]] = duplicate_removed

    payload = {
        "metadata": diagnostics,
        "sourceSheetId": SOURCE_SHEET_ID,
        "publicPostsJson": str(PUBLIC_POSTS_JSON),
        "rankJson": str(RANK_JSON) if RANK_JSON else "",
        "columns": OUTPUT_COLUMNS,
        "sort": "키워드 오름차순, 같은 키워드 안 발행일 최신순, 동일키워드발행수 1 포함",
        "results": results,
        "summary": [
            {
                "sheetName": blog["sheetName"],
                "blogId": blog["blogId"],
                "rows": len(blog["rows"]),
            }
            for blog in results
        ],
    }

    json_path = OUT_DIR / f"{OUT_BASE}.json"
    xlsx_path = OUT_DIR / f"{OUT_BASE}.xlsx"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(results, xlsx_path)

    for blog in results:
        write_tsv(blog, TSV_DIR / f"{blog['sheetName']}.tsv")

    print(json.dumps({"json": str(json_path), "xlsx": str(xlsx_path), "tsvDir": str(TSV_DIR), "summary": payload["summary"], "metadata": diagnostics}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
